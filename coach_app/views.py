from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.utils import timezone
from django import forms
from .models import Coach, Location, Student, Attendance
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO


# Forms
class LocationForm(forms.ModelForm):
    class Meta:
        model = Location
        fields = ['name', 'time_slot']

class StudentForm(forms.ModelForm):
    class Meta:
        model = Student
        fields = ['name', 'age', 'blood_group', 'fees_paid']


# Login & Registration
def coach_login(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            try:
                coach = user.coach
                login(request, user)
                return redirect('select_location', coach_id=coach.id)
            except Coach.DoesNotExist:
                messages.error(request, "This account is not linked to any coach.")
        else:
            messages.error(request, "Invalid username or password.")
        
        return redirect('coach_login')  
    
    return render(request, 'login.html')


@login_required
def coach_logout(request):
    from django.contrib.messages import get_messages
    list(get_messages(request))
    logout(request)
    messages.success(request, "Logged out successfully!")
    return redirect('coach_login')

def register_coach(request):
    if request.method == "POST":
        name = request.POST['name']
        username = request.POST['username']
        password = request.POST['password']
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already taken!')
            return render(request, 'register.html')
        user = User.objects.create_user(username=username, password=password)
        Coach.objects.create(name=name, user=user)
        messages.success(request, f"Welcome {name}! Account created successfully.")
        return redirect('coach_login')
    return render(request, 'register.html')


# Main Views
@login_required
def select_location(request, coach_id):
    coach = get_object_or_404(Coach, id=coach_id, user=request.user)
    locations = coach.locations.all()
    return render(request, 'select_location.html', {
        'coach': coach,
        'locations': locations
    })


@login_required
def attendance(request, location_id):
    location = get_object_or_404(Location, id=location_id, coach__user=request.user)
    students = location.students.all().order_by('name')
    today = date.today()

    # FIXED: Get present students using student__location instead of location field
    present_student_ids = Attendance.objects.filter(
        student__location=location,   # This works because Student has location ForeignKey
        date=today,
        present=True
    ).values_list('student_id', flat=True)

    if request.method == 'POST':
        for student in students:
            is_present = request.POST.get(f'present_{student.id}') == 'on'
            Attendance.objects.update_or_create(
                student=student,
                date=today,
                defaults={'present': is_present}
            )
        messages.success(request, 'Attendance saved successfully!')
        return redirect('attendance', location_id=location_id)

    context = {
        'location': location,
        'students': students,
        'present_student_ids': present_student_ids,
        'today': today,
    }
    return render(request, 'attendance.html', context)

@login_required
def student_detail(request, student_id):
    student = get_object_or_404(Student, id=student_id, location__coach__user=request.user)
    return render(request, 'student_detail.html', {'student': student})


# Add/Edit Student
@login_required
def add_student(request, location_id):
    location = get_object_or_404(Location, id=location_id, coach__user=request.user)
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            student = form.save(commit=False)
            student.location = location
            student.save()
            messages.success(request, f"{student.name} added successfully!")
            return redirect('attendance', location_id=location.id)
    else:
        form = StudentForm()
    
    return render(request, 'form.html', {
        'form': form,
        'title': 'Add New Student',
        'location_id': location.id
    })


@login_required
def edit_student(request, student_id):
    student = get_object_or_404(Student, id=student_id, location__coach__user=request.user)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, f"{student.name} updated successfully!")
            return redirect('attendance', location_id=student.location.id)
    else:
        form = StudentForm(instance=student)
    
    return render(request, 'form.html', {
        'form': form,
        'title': 'Edit Student',
        'student': student,
        'location_id': student.location.id
    })


# Location Management
@login_required
def add_location(request, coach_id):
    coach = get_object_or_404(Coach, id=coach_id, user=request.user)
    if request.method == 'POST':
        form = LocationForm(request.POST)
        if form.is_valid():
            location = form.save(commit=False)
            location.coach = coach
            location.save()
            messages.success(request, f"Batch '{location.name}' created!")
            return redirect('select_location', coach_id=coach.id)
    else:
        form = LocationForm()
    return render(request, 'form.html', {
        'form': form,
        'title': 'Add New Batch',
        'location_id': None
    })


@login_required
def edit_location(request, location_id):
    location = get_object_or_404(Location, id=location_id, coach__user=request.user)
    if request.method == 'POST':
        form = LocationForm(request.POST, instance=location)
        if form.is_valid():
            form.save()
            messages.success(request, "Batch updated!")
            return redirect('select_location', coach_id=location.coach.id)
    else:
        form = LocationForm(instance=location)
    return render(request, 'form.html', {
        'form': form,
        'title': 'Edit Batch',
        'location_id': location.id
    })


@login_required
def delete_location(request, location_id):
    location = get_object_or_404(Location, id=location_id, coach__user=request.user)
    coach_id = location.coach.id
    location.delete()
    messages.success(request, "Batch deleted.")
    return redirect('select_location', coach_id=coach_id)


@login_required
def delete_student(request, student_id):
    student = get_object_or_404(Student, id=student_id, location__coach__user=request.user)
    location_id = student.location.id
    student.delete()
    messages.success(request, f"{student.name} removed.")
    return redirect('attendance', location_id=location_id)


# Reports
@login_required
def attendance_history(request, location_id):
    location = get_object_or_404(Location, id=location_id, coach__user=request.user)
    records = Attendance.objects.filter(student__location=location).order_by('-date', 'student__name')
    
    if request.GET.get('export') == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance_{location.name.replace(' ', '_')}"

        headers = ['Date', 'Student Name', 'Age', 'Blood Group', 'Fees Paid', 'Status']
        ws.append(headers)
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4299E1", end_color="4299E1", fill_type="solid")

        for record in records:
            student = record.student
            ws.append([
                record.date.strftime('%d-%m-%Y'),
                student.name,
                student.age,
                student.blood_group or "-",
                "YES" if student.fees_paid else "NO",
                "Present" if record.present else "Absent"
            ])
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Attendance_{location.name}_{timezone.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response

    return render(request, 'history.html', {
        'location': location,
        'records': records
    })


@login_required
def fees_pdf_report(request, location_id):
    location = get_object_or_404(Location, id=location_id, coach__user=request.user)
    students = location.students.all()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Fees_Report_{location.name}_{timezone.now().strftime("%Y%m%d")}.pdf"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph(f"<font size='18'><b>Fees Report</b></font><br/>{location.name} • {location.time_slot}", styles['Title']))
    elements.append(Paragraph(f"Generated on: {timezone.now().strftime('%d %B %Y')}<br/><br/>", styles['Normal']))

    data = [['No.', 'Student Name', 'Age', 'Fees Paid']]
    for i, s in enumerate(students, 1):
        data.append([str(i), s.name, str(s.age), 'YES' if s.fees_paid else 'NO'])

    table = Table(data, colWidths=[40, 200, 60, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4299E1')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 1, colors.lightgrey),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f7fafc')),
    ]))
    elements.append(table)
    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


def intro(request):
    return render(request, 'intro.html')